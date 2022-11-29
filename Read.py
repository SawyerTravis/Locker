#!/usr/bin/env python

from time import sleep
import sys
import openpyxl
import RPi.GPIO as GPIO
from mfrc522 import SimpleMFRC522

NUMLOCKERS = 2

reader = SimpleMFRC522()

# Locker Workbook is opened
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("/home/pi/MFRC522-python/locker.xlsx")

sheet = wb.active

# GPIO Setup
lockerBase = 36
for i in range(0, NUMLOCKERS) :
    GPIO.setup(lockerBase + i, GPIO.OUT)

while (True) :
    #id = 207990694991
    #text = nothing
        id, text = reader.read()
        print(id)
        print(text)
        for i in range(1, NUMLOCKERS + 1):
            print(i)
            print(sheet.cell(i, 1).value)
            if sheet.cell(i, 1).value == id:
                print("Opening locker " + str(i))
                GPIO.output(lockerBase+i-1, GPIO.HIGH)
                sleep(1)
                GPIO.output(lockerBase+i-1, GPIO.LOW)
                break