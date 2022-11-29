# Python program to read an excel file
 
# import openpyxl module
import openpyxl
 
# Give the location of the file
path = "/home/pi/MFRC522-python/locker.xlsx"
wb = openpyxl.Workbook()
wb.save(path)
 
# Get workbook active sheet object
# from the active attribute
sheet = wb.active
 
# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell() method.
cell1 = sheet.cell(row = 1, column = 1)
cell1.value = 972251405331
cell1 = sheet.cell(row = 2, column = 1)
cell1.value = 634378661861
 
# Print value of cell object
# using the value attribute
print(cell1.value)
wb.save("locker.xlsx")